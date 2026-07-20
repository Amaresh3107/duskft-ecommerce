from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional
import csv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from bson import ObjectId
from database import db
from models import Product
from deps import require_roles, get_optional_session
from business import slugify, calculate_price, now_iso

router = APIRouter(prefix='/api/products', tags=['products'])

IMPORT_HEADERS = [
    'sku', 'name', 'categoryName', 'description', 'basePrice', 'moq', 'totalStock', 'stock',
    'colors', 'sizes', 'tierPricing', 'videoUrl', 'images', 'status',
]


@router.get('/import/template')
async def download_import_template(session: dict = Depends(require_roles('admin'))):
    """A downloadable .xlsx admins can fill in and re-upload via /import.
    Header row is frozen (stays visible while scrolling) and locked (sheet
    protection is on, so the column names can't be accidentally overwritten)
    — data rows are explicitly left unlocked so they're still editable."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    ws.append(IMPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        cell.protection = Protection(locked=True)
    ws.freeze_panes = 'A2'

    ws.append([
        'DRS-002', 'Floral Maxi Dress', 'Dresses & Sets',
        'Lightweight floral print maxi dress, breathable fabric',
        850, 15, 200, 200,
        'Red,Blue,Green', 'S,M,L,XL',
        '15:850,25:800,50:750', '',
        'https://example.com/image1.jpg,https://example.com/image2.jpg',
        'active',
    ])
    ws.append(['# categoryName must match an existing category name exactly (case-insensitive), or leave blank'])
    ws.append(['# tierPricing format: minQty:price pairs separated by commas, e.g. 15:850,25:800 — leave blank for none'])
    ws.append(["# colors / sizes / images: comma-separated. images must be full URLs — local upload from a spreadsheet isn't supported yet"])

    # Data rows (including a good chunk of blank ones below, for typing new
    # rows) stay editable even with sheet protection turned on.
    for row in ws.iter_rows(min_row=2, max_row=500, max_col=len(IMPORT_HEADERS)):
        for cell in row:
            cell.protection = Protection(locked=False)

    for i, header in enumerate(IMPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 4)

    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="product_import_template.xlsx"'},
    )


def _rows_from_csv(raw: bytes):
    try:
        text = raw.decode('utf-8-sig')  # utf-8-sig strips Excel's BOM if present
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail='Could not read this file as text — please save it as CSV (UTF-8).')
    return list(csv.DictReader(io.StringIO(text)))


def _rows_from_xlsx(raw: bytes):
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail='Could not read this .xlsx file — is it a valid Excel file?')
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        row = {headers[i]: ('' if values[i] is None else str(values[i])) for i in range(min(len(headers), len(values)))}
        rows.append(row)
    return rows


def _parse_and_build_product(row: dict, category_lookup: dict) -> Product:
    name = (row.get('name') or '').strip()
    sku = (row.get('sku') or '').strip()
    base_price = float(row.get('basePrice') or 0)
    moq = int(float(row.get('moq') or 1))
    total_stock = int(float(row.get('totalStock') or 0))
    stock = int(float(row.get('stock') or total_stock))
    if stock > total_stock:
        raise ValueError('stock cannot exceed totalStock')

    colors = [c.strip() for c in (row.get('colors') or '').split(',') if c.strip()]
    sizes = [s.strip() for s in (row.get('sizes') or '').split(',') if s.strip()]
    images = [u.strip() for u in (row.get('images') or '').split(',') if u.strip()]

    category_name = (row.get('categoryName') or '').strip().lower()
    category_id = category_lookup.get(category_name) if category_name else None
    if category_name and not category_id:
        raise ValueError(f'No category named "{row.get("categoryName")}" — check spelling or leave blank')

    tier_pricing = []
    tier_str = (row.get('tierPricing') or '').strip()
    if tier_str:
        for pair in tier_str.split(','):
            pair = pair.strip()
            if not pair:
                continue
            if ':' not in pair:
                raise ValueError(f'Bad tierPricing entry "{pair}" — expected format minQty:price')
            min_qty_str, price_str = pair.split(':', 1)
            tier_pricing.append({'minQty': int(float(min_qty_str)), 'price': float(price_str)})

    return Product(
        sku=sku, name=name, slug=slugify(name), categoryId=category_id,
        description=(row.get('description') or '').strip(),
        images=images, videoUrl=(row.get('videoUrl') or '').strip(),
        colors=colors, sizes=sizes, tierPricing=tier_pricing,
        basePrice=base_price, moq=moq, totalStock=total_stock, stock=stock,
        status=(row.get('status') or 'active').strip().lower() or 'active',
        createdAt=now_iso(),
    )


@router.post('/import')
async def bulk_import_products(file: UploadFile = File(...), session: dict = Depends(require_roles('admin'))):
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail='Please upload a .csv or .xlsx file.')

    raw = await file.read()
    rows = _rows_from_xlsx(raw) if filename.endswith('.xlsx') else _rows_from_csv(raw)

    category_docs = await db.categories.find().to_list(1000)
    category_lookup = {c['name'].strip().lower(): str(c['_id']) for c in category_docs}

    created = 0
    errors = []
    row_num = 1  # header is row 1

    for row in rows:
        row_num += 1
        name = (row.get('name') or '').strip()
        if not name or name.startswith('#'):
            continue  # skip blank rows and the "# ..." instruction rows from the template

        try:
            product = _parse_and_build_product(row, category_lookup)
            await db.products.insert_one(product.to_mongo())
            created += 1
        except Exception as e:
            errors.append({'row': row_num, 'product': name, 'error': str(e)})

    return {'created': created, 'errors': errors}


@router.get('')
async def list_products(categoryId: Optional[str] = None, search: Optional[str] = None,
                         includeInactive: bool = False, session: dict = Depends(get_optional_session)):
    query = {}
    if includeInactive and session and session['role'] in ('admin', 'staff'):
        pass
    else:
        query['status'] = 'active'
    if categoryId:
        query['categoryId'] = categoryId
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'sku': {'$regex': search, '$options': 'i'}},
        ]
    docs = await db.products.find(query).to_list(1000)
    return [Product.from_mongo(d) for d in docs]


@router.get('/{product_id}')
async def get_product(product_id: str):
    doc = await db.products.find_one({'_id': ObjectId(product_id)})
    if not doc:
        raise HTTPException(status_code=404, detail='Product not found.')
    return Product.from_mongo(doc)


@router.post('')
async def create_product(payload: dict, session: dict = Depends(require_roles('admin', 'staff'))):
    total_stock = int(payload.get('totalStock', payload.get('stock', 0)))
    stock = int(payload.get('stock', payload.get('totalStock', 0)))
    if stock > total_stock:
        raise HTTPException(status_code=400, detail='Current Stock cannot exceed Total Stock.')
    product = Product(
        sku=payload.get('sku', ''),
        name=payload['name'],
        slug=payload.get('slug') or slugify(payload.get('name', '')),
        categoryId=payload.get('categoryId'),
        description=payload.get('description', ''),
        images=payload.get('images', []),
        videoUrl=payload.get('videoUrl', ''),
        colors=payload.get('colors', []),
        sizes=payload.get('sizes', []),
        tierPricing=payload.get('tierPricing', []),
        basePrice=float(payload.get('basePrice', 0)),
        moq=int(payload.get('moq', 1)),
        totalStock=total_stock,
        stock=stock,
        status=payload.get('status', 'active'),
        createdAt=now_iso(),
    )
    result = await db.products.insert_one(product.to_mongo())
    doc = await db.products.find_one({'_id': result.inserted_id})
    return Product.from_mongo(doc)


@router.put('/{product_id}')
async def update_product(product_id: str, payload: dict, session: dict = Depends(require_roles('admin', 'staff'))):
    existing = await db.products.find_one({'_id': ObjectId(product_id)})
    if not existing:
        raise HTTPException(status_code=404, detail='Product not found.')

    if 'totalStock' in payload or 'stock' in payload:
        effective_total = int(payload.get('totalStock', existing.get('totalStock', 0)))
        effective_stock = int(payload.get('stock', existing.get('stock', 0)))
        if effective_stock > effective_total:
            raise HTTPException(status_code=400, detail='Current Stock cannot exceed Total Stock.')

    updates = {k: v for k, v in payload.items() if k not in ('id', '_id')}
    updates['updatedAt'] = now_iso()
    await db.products.update_one({'_id': ObjectId(product_id)}, {'$set': updates})
    doc = await db.products.find_one({'_id': ObjectId(product_id)})
    return Product.from_mongo(doc)


@router.delete('/{product_id}')
async def delete_product(product_id: str, session: dict = Depends(require_roles('admin'))):
    result = await db.products.delete_one({'_id': ObjectId(product_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Product not found.')
    return {'success': True}


@router.post('/calc-price')
async def calc_price(payload: dict):
    product_id = payload.get('productId')
    quantity = int(payload.get('quantity', 0))
    doc = await db.products.find_one({'_id': ObjectId(product_id)})
    if not doc:
        raise HTTPException(status_code=404, detail='Product not found.')
    result = calculate_price(doc, quantity)
    result['productId'] = product_id
    result['quantity'] = quantity
    return result
